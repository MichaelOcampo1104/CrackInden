import pandas as pd
import cv2
import numpy as np
import fitz  # PyMuPDF
import io
import zipfile
import os
import re
from openpyxl import Workbook
from PIL import Image as PilImage
from openpyxl.drawing.image import Image as OpenpyxlImage


# Define the PDF path
pdf_path = r'C:\Users\hp\Documents\Python Scripts\crackinden\Sample3.pdf'

# Define the directory to save images
image_dir = r'C:\Users\hp\Documents\Python Scripts\crackinden\images'

# Check if the directory exists; if not, create it
if not os.path.exists(image_dir):
    os.makedirs(image_dir)

# Initialize a dictionary to store the text and image references of filtered pages
filtered_data_xlsx = {
    'Page Number': [],
    'Full Text': [],
    'Image Reference': []
}

# Loop through each page, extract text, and check for images
pdf_document = fitz.open(pdf_path)
for page_num in range(len(pdf_document)):
    page = pdf_document.load_page(page_num)
    page_text = page.get_text().lower()  # Convert to lowercase for case-insensitive search

    # Check if the page contains any of the terms "crack," "damage," or "defect"
    if any(term in page_text for term in ['crack', 'damage', 'defect']):
        filtered_data_xlsx['Page Number'].append(page_num + 1)  # Page numbering starts from 1
        filtered_data_xlsx['Full Text'].append(page.get_text())

        image_list = page.get_images(full=True)
        image_ref = []
        for img_index, img in enumerate(image_list):
            xref = img[0]
            base_image = pdf_document.extract_image(xref)
            image_bytes = base_image["image"]
            
            # Convert to a numpy array and save as an image
            image_np = np.array(PilImage.open(io.BytesIO(image_bytes)))
            image_filename = f"Page_{page_num + 1}_Image_{img_index + 1}.png"
            image_filepath = os.path.join(image_dir, image_filename)
            cv2.imwrite(image_filepath, cv2.cvtColor(image_np, cv2.COLOR_RGB2BGR))
            
            image_ref.append(image_filename)
        filtered_data_xlsx['Image Reference'].append(", ".join(image_ref))

# Create a DataFrame from the dictionary
filtered_data_df = pd.DataFrame(filtered_data_xlsx)

# Remove all kinds of whitespaces from "Full Text" column
filtered_data_df['Full Text'] = filtered_data_df['Full Text'].apply(lambda x: re.sub(r'\s+', ' ', str(x)).strip())

# Create a new Excel workbook and select the active worksheet
wb = Workbook()
ws1 = wb.active
ws1.title = "Filtered Data"

# Write headers to the Excel file
ws1.append(['Page Number', 'Full Text', 'Image Reference'])

row_num = 2  # Initialize row number; Excel row numbers start from 1, and we have a header row

# Loop through the filtered data to populate the Excel file
for page_num, full_text, image_refs in zip(filtered_data_xlsx['Page Number'], filtered_data_xlsx['Full Text'], filtered_data_xlsx['Image Reference']):
    # Remove all kinds of whitespaces from "Full Text"
    cleaned_full_text = re.sub(r'\s+', ' ', full_text).strip()

    # Write data to Excel
    ws1.append([page_num, cleaned_full_text, image_refs])
    
    # Embed images into respective columns starting from Column C (if any)
    col_num = 3  # Initialize column number for images; A=1, B=2, C=3, ...
    if image_refs:
        for img_index, img_file in enumerate(image_refs.split(", ")):
            img_path = os.path.join(image_dir, img_file.strip())
            if os.path.exists(img_path):
                img = OpenpyxlImage(img_path)
                img.width = 60
                img.height = 80
                cell_num = ws1.cell(row=row_num, column=col_num).coordinate
                ws1.column_dimensions[chr(64 + col_num)].width = 15
                ws1.row_dimensions[row_num].height = 90
                ws1.add_image(img, cell_num)
                
                # Increment column number for the next image
                col_num += 1

    # Increment row number for the next page
    row_num += 1


# Save the DataFrame to a new sheet in the same Excel workbook
ws2 = wb.create_sheet(title='DataFrame')
for r_idx, row in enumerate(filtered_data_df.values, 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx)
        cell.value = value

            
# Save the DataFrame to an Excel file
xlsx_path = r'C:\Users\hp\Documents\Python Scripts\crackinden\file_filtered_with_images.xlsx'
wb.save(xlsx_path)
#This code is working





# Create a ZIP file to store the images
zip_path = r'C:\Users\hp\Documents\Python Scripts\crackinden\file_filtered.zip'
with zipfile.ZipFile(zip_path, 'w') as zipf:
    if os.path.exists(image_dir):
        for image_file in os.listdir(image_dir):
            zipf.write(os.path.join(image_dir, image_file), image_file)
            

