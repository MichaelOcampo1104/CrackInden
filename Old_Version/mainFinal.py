import argparse
import pandas as pd
import cv2
import numpy as np
import fitz  # PyMuPDF
import io
import zipfile
import os
import re
from openpyxl import Workbook
from PIL import Image as PilImage
from openpyxl.drawing.image import Image as OpenpyxlImage

def main():
    parser = argparse.ArgumentParser(description='PDF Image Extractor')
    parser.add_argument('--input', type=str, help='Path to input PDF file')
    parser.add_argument('--output', type=str, help='Path to output Excel file')
    args = parser.parse_args()

    if not (args.input and args.output):
        print('Please provide both input and output paths.')
        return

    pdf_path = args.input
    excel_path = args.output
    image_dir = os.path.dirname(excel_path)

    if not os.path.exists(image_dir):
        os.makedirs(image_dir)

    filtered_data_xlsx = {
        'Page Number': [],
        'Full Text': [],
        'Image Reference': []
    }

    pdf_document = fitz.open(pdf_path)
    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        page_text = page.get_text().lower()

        if any(term in page_text for term in ['crack', 'damage', 'defect']):
            filtered_data_xlsx['Page Number'].append(page_num + 1)
            filtered_data_xlsx['Full Text'].append(page.get_text())

            image_list = page.get_images(full=True)
            image_ref = []
            for img_index, img in enumerate(image_list):
                xref = img[0]
                base_image = pdf_document.extract_image(xref)
                image_bytes = base_image['image']

                image_np = np.array(PilImage.open(io.BytesIO(image_bytes)))
                image_filename = f'Page_{page_num + 1}_Image_{img_index + 1}.png'
                image_filepath = os.path.join(image_dir, image_filename)
                cv2.imwrite(image_filepath, cv2.cvtColor(image_np, cv2.COLOR_RGB2BGR))

                image_ref.append(image_filename)
            filtered_data_xlsx['Image Reference'].append(', '.join(image_ref))

    filtered_data_df = pd.DataFrame(filtered_data_xlsx)
    filtered_data_df['Full Text'] = filtered_data_df['Full Text'].apply(lambda x: re.sub(r'\s+', ' ', str(x)).strip())

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Filtered Data'

    ws1.append(['Page Number', 'Full Text', 'Image Reference'])

    row_num = 2

    for page_num, full_text, image_refs in zip(filtered_data_xlsx['Page Number'], filtered_data_xlsx['Full Text'], filtered_data_xlsx['Image Reference']):
        cleaned_full_text = re.sub(r'\s+', ' ', full_text).strip()
        ws1.append([page_num, cleaned_full_text, image_refs])

        col_num = 3
        if image_refs:
            for img_index, img_file in enumerate(image_refs.split(', ')):
                img_path = os.path.join(image_dir, img_file.strip())
                if os.path.exists(img_path):
                    img = OpenpyxlImage(img_path)
                    img.width = 60
                    img.height = 80
                    cell_num = ws1.cell(row=row_num, column=col_num).coordinate
                    ws1.column_dimensions[chr(64 + col_num)].width = 15
                    ws1.row_dimensions[row_num].height = 90
                    ws1.add_image(img, cell_num)
                    col_num += 1

        row_num += 1

    ws2 = wb.create_sheet(title='DataFrame')
    for r_idx, row in enumerate(filtered_data_df.values, 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx)
            cell.value = value

    wb.save(excel_path)

if __name__ == '__main__':
    main()